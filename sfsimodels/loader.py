__author__ = 'maximmillen'

import numpy as np
import openpyxl


def generate_new_list(fp):
    wb2 = openpyxl.load_workbook(fp)
    xi = wb2.get_sheet_by_name('Inputs')

    i = 1
    for row in xi.rows:
        val = row[0].value

        if val is not None:
            parts = val.split(" [")
            val = parts[0]

            val = val.replace(" ", "")
            if "#" not in val:
                print('"{0}",'.format(val))
            else:
                print(val)
        i += 1


def generate_new_index(fp):
    wb2 = openpyxl.load_workbook(fp)
    xi = wb2.get_sheet_by_name('Inputs')
    i = 1
    for row in xi.rows:
        val = row[0].value

        # if cell is not None:
        # val = cell.value
        if val is not None:
            parts = val.split(" [")
            val = parts[0]
            units = ""
            if len(parts) == 2:
                units = "  # " + parts[1].replace("]", "")
            val = val.replace(" ", "")
            if "#" not in val:
                print('    indy["{0}"] = {1}{2}'.format(val, i, units))
            else:
                print("    " + val)
        i += 1


def add_inputs_to_object(obj, values):
    """
    A generic function to load object parameters based on a dictionary list.
    :param obj: Object
    :param values: Dictionary
    :return:
    """
    for item in obj.inputs:
        if hasattr(obj, item):
            # print(item)
            setattr(obj, item, values[item])


def load_from_xlsx(sss, fp):
    """
    Add the SoilStructureSystem object properties using a spreadsheet.
    :param sss: SoilStructureSystem object
    :param fp: File path to spreadsheet
    :return:
    """
    try:
        wb2 = openpyxl.load_workbook(fp, data_only=True)
    except FileNotFoundError:
        print("Can not find file: {0}".format(fp))
        raise FileNotFoundError

    try:
        xi = wb2.get_sheet_by_name('Inputs')
    except KeyError:
        print('Sheet name: "Inputs" not found! in {0}'.format(fp))
        raise KeyError

    d_values = {}
    p_values = {}
    for row in xi.rows:
        val = row[0].value
        if val is None:
            continue
        if "#" in val:
            continue
        parts = val.split(" [")
        val = parts[0]
        val = val.replace(" ", "")
        if isinstance(row[1].value, str) and "=" in row[1].value:
            print(row[1].value)
        d_values[val] = row[1].value
        p_values[val] = []
        x = 4
        while row[x].value is not None:
            p_values[val].append(row[x].value)
            x += 1
            if len(row) == x:
                break

    add_inputs_to_object(sss, d_values)
    add_inputs_to_object(sss.sp, d_values)
    add_inputs_to_object(sss.fd, d_values)
    add_inputs_to_object(sss.bd, d_values)
    add_inputs_to_object(sss.hz, d_values)
    return p_values


def load_sample_data(sss):
    """
    Sample data for the SoilStructureSystem object
    :param sss:
    :return:
    """

    load_soil_sample_data(sss.sp)  # soil
    load_foundation_sample_data(sss.fd)  # foundation
    load_structure_sample_data(sss.bd)  # structure
    load_hazard_sample_data(sss.hz)  # hazard


def load_soil_sample_data(sp):
    """
    Sample data for the Soil object
    :param sp: Soil Object
    :return:
    """
    # soil
    sp.g_mod = 60.0e6  # [Pa]
    sp.phi = 30  # [degrees]
    sp.relative_density = .40  # [decimal]
    sp.height_crust = 2.  # [m]
    sp.height_liq = 5.  # [m]
    sp.gwl = 2.  # [m], ground water level
    sp.unit_weight_crust = 17000  # [N/m3]
    sp.unit_sat_weight_liq = 18000  # [N/m3]
    sp.unit_weight_water = 9800  # [N/m3]
    sp.cohesion = 10.0  # [Pa]
    sp.poissons_ratio = 0.22
    sp.e_min = 0.55
    sp.e_max = 0.95
    sp.e_cr0 = 0.79  # Jin et al. 2015
    sp.p_cr0 = 0.7  # Jin et al. 2015
    sp.lamb_crl = 0.015  # Jin et al. 2015


def load_foundation_sample_data(fd):
    """
    Sample data for the Foundation object
    :param fd: Foundation Object
    :return:
    """
    # foundation
    fd.width = 16.0  # m
    fd.length = 18.0  # m
    fd.depth = 0.0  # m


def load_structure_sample_data(st):
    """
    Sample data for the Structure object
    :param st: Structure Object
    :return:
    """
    # structure
    st.h_eff = 9.0  # m
    st.mass_eff = 120e3  # kg
    st.t_eff = 1.0  # s
    st.mass_ratio = 1.0  # ratio of mass acting horizontal to vertically


def load_hazard_sample_data(hz):
    """
    Sample data for the Hazard object
    :param hz: Hazard Object
    :return:
    """
    # hazard
    hz.z_factor = 0.3  # Hazard factor
    hz.r_factor = 1.0  # Return period factor
    hz.n_factor = 1.0  # Near-fault factor
    hz.magnitude = 7.5  # Magnitude of earthquake
    hz.corner_period = 4.0  # s
    hz.corner_acc_factor = 0.55
    return hz


def load_building_sample_data(bd):
    """
    Sample data for the Building object
    :param bd:
    :return:
    """
    number_of_storeys = 6
    interstorey_height = 3.4  # m
    masses = 40.0e3  # kg

    bd.interstorey_heights = interstorey_height * np.ones(number_of_storeys)
    bd.floor_length = 18.0  # m
    bd.floor_width = 16.0  # m
    bd.storey_masses = masses * np.ones(number_of_storeys)  # kg


def load_frame_building_sample_data(fb):
    """
    Sample data for the FrameBuilding object
    :param fb:
    :return:
    """
    load_building_sample_data(fb)

    fb.bay_lengths = [6., 6.0, 6.0]
    fb.beam_depths = [.5]
    fb.n_seismic_frames = 3
    fb.n_gravity_frames = 0


