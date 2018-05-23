import openpyxl
from sfsimodels.loader import add_inputs_to_object

__author__ = 'maximmillen'


def generate_new_list(fp):
    wb2 = openpyxl.load_workbook(fp)
    xi = wb2.get_sheet_by_name('Inputs')

    i = 1
    for row in xi.rows:
        val = row[0].value

        if val is not None:
            parts = val.split(" [")
            val = parts[0]

            val = val.replace(" ", "")
            if "#" not in val:
                print('"{0}",'.format(val))
            else:
                print(val)
        i += 1


def generate_new_index(fp):
    wb2 = openpyxl.load_workbook(fp)
    xi = wb2.get_sheet_by_name('Inputs')
    i = 1
    for row in xi.rows:
        val = row[0].value

        # if cell is not None:
        # val = cell.value
        if val is not None:
            parts = val.split(" [")
            val = parts[0]
            units = ""
            if len(parts) == 2:
                units = "  # " + parts[1].replace("]", "")
            val = val.replace(" ", "")
            if "#" not in val:
                print('    indy["{0}"] = {1}{2}'.format(val, i, units))
            else:
                print("    " + val)
        i += 1


def load_from_xlsx(sss, fp):
    """
    Add the SoilStructureSystem object properties using a spreadsheet.
    :param sss: SoilStructureSystem object
    :param fp: File path to spreadsheet
    :return:
    """
    try:
        wb2 = openpyxl.load_workbook(fp, data_only=True)
    except FileNotFoundError:
        print("Can not find file: {0}".format(fp))
        raise FileNotFoundError

    try:
        xi = wb2.get_sheet_by_name('Inputs')
    except KeyError:
        print('Sheet name: "Inputs" not found! in {0}'.format(fp))
        raise KeyError

    d_values = {}
    p_values = {}
    for row in xi.rows:
        val = row[0].value
        if val is None:
            continue
        if "#" in val:
            continue
        parts = val.split(" [")
        val = parts[0]
        val = val.replace(" ", "")
        # if isinstance(row[1].value, str) and "=" in row[1].value:
        #     print(row[1].value)
        d_values[val] = row[1].value
        p_values[val] = []
        x = 4
        while row[x].value is not None:
            p_values[val].append(row[x].value)
            x += 1
            if len(row) == x:
                break

    add_inputs_to_object(sss, d_values)
    add_inputs_to_object(sss.sp, d_values)
    add_inputs_to_object(sss.fd, d_values)
    add_inputs_to_object(sss.bd, d_values)
    add_inputs_to_object(sss.hz, d_values)
    return p_values